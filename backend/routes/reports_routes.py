from fastapi import APIRouter, HTTPException, Depends, Query, Response
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime, timezone, timedelta
from motor.motor_asyncio import AsyncIOMotorClient
import os
from middleware.auth import get_current_user, require_role
from models.user import UserRole
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

router = APIRouter(prefix="/reports", tags=["Reports"])

# MongoDB connection
MONGO_URL = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
client = AsyncIOMotorClient(MONGO_URL)
db = client.distribution_db

@router.get("/sales/export")
async def export_sales_report(
    format: str = Query("xlsx", regex="^(xlsx|pdf)$"),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.ACCOUNTING]))
):
    """Export sales report in Excel or PDF format"""
    
    # Calculate date range
    now = datetime.now(timezone.utc)
    if start_date and end_date:
        start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    else:
        start = now - timedelta(days=30)
        end = now
    
    # Get orders
    orders = await db.orders.find({
        "created_at": {"$gte": start, "$lte": end},
        "status": {"$in": ["delivered", "dispatched", "ready", "preparing", "approved"]}
    }).to_list(10000)
    
    # Enrich with customer and product details
    report_data = []
    for order in orders:
        customer = await db.users.find_one({"id": order.get('customer_id')})
        customer_name = customer.get('full_name', 'Unknown') if customer else 'Unknown'
        
        for product in order.get('products', []):
            prod = await db.products.find_one({"id": product.get('product_id')})
            product_name = prod.get('name', 'Unknown') if prod else 'Unknown'
            
            report_data.append({
                "order_number": order.get('order_number'),
                "date": order.get('created_at').strftime('%Y-%m-%d %H:%M'),
                "customer_name": customer_name,
                "product_name": product_name,
                "quantity": product.get('quantity', 0),
                "price": product.get('price', 0),
                "total": product.get('price', 0) * product.get('quantity', 0),
                "status": order.get('status')
            })
    
    if format == "xlsx":
        return generate_sales_excel(report_data, start, end)
    else:
        return generate_sales_pdf(report_data, start, end)

@router.get("/stock/export")
async def export_stock_report(
    format: str = Query("xlsx", regex="^(xlsx|pdf)$"),
    warehouse_id: Optional[str] = None,
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.WAREHOUSE_MANAGER]))
):
    """Export stock report in Excel or PDF format"""
    
    # Get inventory data
    query = {}
    if warehouse_id:
        query['warehouse_id'] = warehouse_id
    
    inventory_items = await db.inventory.find(query).to_list(10000)
    
    report_data = []
    for item in inventory_items:
        product = await db.products.find_one({"id": item.get('product_id')})
        warehouse = await db.warehouses.find_one({"id": item.get('warehouse_id')})
        
        report_data.append({
            "product_name": product.get('name', 'Unknown') if product else 'Unknown',
            "product_sku": product.get('sku', 'N/A') if product else 'N/A',
            "warehouse_name": warehouse.get('name', 'Unknown') if warehouse else 'Unknown',
            "total_units": item.get('total_units', 0),
            "status": "Out of Stock" if item.get('is_out_of_stock') else "In Stock",
            "last_supply_date": item.get('last_supply_date').strftime('%Y-%m-%d') if item.get('last_supply_date') else 'N/A'
        })
    
    if format == "xlsx":
        return generate_stock_excel(report_data)
    else:
        return generate_stock_pdf(report_data)

@router.get("/sales-agents/export")
async def export_sales_agents_report(
    format: str = Query("xlsx", regex="^(xlsx|pdf)$"),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(require_role([UserRole.ADMIN]))
):
    """Export sales agents performance report"""
    
    # Calculate date range
    now = datetime.now(timezone.utc)
    if start_date and end_date:
        start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    else:
        start = now - timedelta(days=30)
        end = now
    
    # Get all sales agents
    sales_agents = await db.users.find({"role": "sales_agent", "is_active": True}).to_list(1000)
    
    report_data = []
    for agent in sales_agents:
        agent_id = agent.get('id')
        
        # Get orders for this agent
        agent_orders = await db.orders.find({
            "sales_rep_id": agent_id,
            "created_at": {"$gte": start, "$lte": end},
            "status": {"$in": ["delivered", "dispatched", "ready", "preparing", "approved"]}
        }).to_list(10000)
        
        total_sales = sum(order.get('total_amount', 0) for order in agent_orders)
        total_orders = len(agent_orders)
        
        # Get customer count
        customers = await db.sales_routes.count_documents({"sales_rep_id": agent_id})
        
        report_data.append({
            "agent_name": agent.get('full_name', agent.get('username')),
            "total_customers": customers,
            "total_orders": total_orders,
            "total_sales": round(total_sales, 2),
            "average_order_value": round(total_sales / total_orders, 2) if total_orders > 0 else 0
        })
    
    report_data.sort(key=lambda x: x['total_sales'], reverse=True)
    
    if format == "xlsx":
        return generate_agents_excel(report_data, start, end)
    else:
        return generate_agents_pdf(report_data, start, end)

@router.get("/logistics/export")
async def export_logistics_report(
    format: str = Query("xlsx", regex="^(xlsx|pdf)$"),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.WAREHOUSE_MANAGER]))
):
    """Export logistics/delivery report"""
    
    # Calculate date range
    now = datetime.now(timezone.utc)
    if start_date and end_date:
        start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    else:
        start = now - timedelta(days=30)
        end = now
    
    # Get delivered orders
    orders = await db.orders.find({
        "delivered_date": {"$gte": start, "$lte": end},
        "status": "delivered"
    }).to_list(10000)
    
    report_data = []
    for order in orders:
        customer = await db.users.find_one({"id": order.get('customer_id')})
        agent = await db.users.find_one({"id": order.get('sales_rep_id')})
        
        # Calculate delivery time
        created = order.get('created_at')
        delivered = order.get('delivered_date')
        delivery_days = (delivered - created).days if created and delivered else 0
        
        report_data.append({
            "order_number": order.get('order_number'),
            "customer_name": customer.get('full_name', 'Unknown') if customer else 'Unknown',
            "agent_name": agent.get('full_name', 'Unknown') if agent else 'Unknown',
            "order_date": created.strftime('%Y-%m-%d') if created else 'N/A',
            "delivery_date": delivered.strftime('%Y-%m-%d') if delivered else 'N/A',
            "delivery_days": delivery_days,
            "total_amount": order.get('total_amount', 0),
            "channel_type": order.get('channel_type', 'N/A')
        })
    
    if format == "xlsx":
        return generate_logistics_excel(report_data, start, end)
    else:
        return generate_logistics_pdf(report_data, start, end)

# Excel generation functions
def generate_sales_excel(data, start_date, end_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Sales Report ({start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')})"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["Order Number", "Date", "Customer", "Product", "Quantity", "Price", "Total", "Status"]
    ws.append([])
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for item in data:
        ws.append([
            item['order_number'],
            item['date'],
            item['customer_name'],
            item['product_name'],
            item['quantity'],
            item['price'],
            item['total'],
            item['status']
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=sales_report_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

def generate_stock_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Report"
    
    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "Stock Report"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["Product Name", "SKU", "Warehouse", "Total Units", "Status", "Last Supply Date"]
    ws.append([])
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for item in data:
        ws.append([
            item['product_name'],
            item['product_sku'],
            item['warehouse_name'],
            item['total_units'],
            item['status'],
            item['last_supply_date']
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=stock_report_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

def generate_agents_excel(data, start_date, end_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Agents Report"
    
    # Title
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = f"Sales Agents Performance Report ({start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')})"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["Agent Name", "Total Customers", "Total Orders", "Total Sales", "Avg Order Value"]
    ws.append([])
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for item in data:
        ws.append([
            item['agent_name'],
            item['total_customers'],
            item['total_orders'],
            item['total_sales'],
            item['average_order_value']
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=sales_agents_report_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

def generate_logistics_excel(data, start_date, end_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logistics Report"
    
    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Logistics/Delivery Report ({start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')})"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["Order Number", "Customer", "Agent", "Order Date", "Delivery Date", "Delivery Days", "Amount", "Channel"]
    ws.append([])
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for item in data:
        ws.append([
            item['order_number'],
            item['customer_name'],
            item['agent_name'],
            item['order_date'],
            item['delivery_date'],
            item['delivery_days'],
            item['total_amount'],
            item['channel_type']
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=logistics_report_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

# PDF generation functions (simplified versions)
def generate_sales_pdf(data, start_date, end_date):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER)
    title = Paragraph(f"Sales Report<br/>({start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')})", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Table
    table_data = [["Order", "Date", "Customer", "Product", "Qty", "Price", "Total", "Status"]]
    for item in data[:50]:  # Limit to 50 rows for PDF
        table_data.append([
            item['order_number'][:15],
            item['date'][:16],
            item['customer_name'][:20],
            item['product_name'][:25],
            str(item['quantity']),
            f"{item['price']:.2f}",
            f"{item['total']:.2f}",
            item['status'][:10]
        ])
    
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=sales_report_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )

def generate_stock_pdf(data):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER)
    title = Paragraph("Stock Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    table_data = [["Product", "SKU", "Warehouse", "Units", "Status"]]
    for item in data[:50]:
        table_data.append([
            item['product_name'][:30],
            item['product_sku'][:15],
            item['warehouse_name'][:20],
            str(item['total_units']),
            item['status']
        ])
    
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=stock_report_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )

def generate_agents_pdf(data, start_date, end_date):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER)
    title = Paragraph(f"Sales Agents Performance Report<br/>({start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')})", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    table_data = [["Agent", "Customers", "Orders", "Sales", "Avg Order"]]
    for item in data:
        table_data.append([
            item['agent_name'][:30],
            str(item['total_customers']),
            str(item['total_orders']),
            f"{item['total_sales']:.2f}",
            f"{item['average_order_value']:.2f}"
        ])
    
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=sales_agents_report_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )

def generate_logistics_pdf(data, start_date, end_date):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER)
    title = Paragraph(f"Logistics/Delivery Report<br/>({start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')})", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    table_data = [["Order", "Customer", "Agent", "Order Date", "Delivery", "Days", "Amount"]]
    for item in data[:50]:
        table_data.append([
            item['order_number'][:15],
            item['customer_name'][:20],
            item['agent_name'][:20],
            item['order_date'],
            item['delivery_date'],
            str(item['delivery_days']),
            f"{item['total_amount']:.2f}"
        ])
    
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=logistics_report_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )
